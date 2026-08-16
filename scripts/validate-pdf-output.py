from pathlib import Path

import pdfplumber
from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "tmp" / "pdfs"

EXPECTED = {
    "quotation-sample.pdf": ("IME-26-Q0099", "TOTAL HT"),
    "proforma-sample.pdf": ("IME-26-F0099", "SOLDE À PAYER"),
    "transmittal-sample.pdf": ("TR-2026-0099", "TRANSMITTED DOCUMENTS"),
    "commercial-invoice-sample.pdf": ("CI-IME-26-F0099", "COMMERCIAL INVOICE"),
    "packing-list-sample.pdf": ("PL-IME-26-F0099", "PACKING LIST"),
    "delivery-note-sample.pdf": ("DN-IME-26-F0099", "DELIVERY NOTE"),
    "calculator-sample.pdf": ("RÉSULTATS TECHNIQUES", "Convertisseur de fréquence"),
    "reports-sample.pdf": ("Rapports & Performance", "Dettes fournisseurs"),
}

PORTRAIT_A4 = (595.28, 841.89)
LANDSCAPE_A4 = (841.89, 595.28)


def close(a: float, b: float, tolerance: float = 2.0) -> bool:
    return abs(a - b) <= tolerance


for filename, required_text in EXPECTED.items():
    path = OUTPUT / filename
    if not path.exists() or path.stat().st_size < 5_000:
        raise AssertionError(f"PDF absent ou anormalement petit : {filename}")

    reader = PdfReader(path)
    if not reader.pages:
        raise AssertionError(f"PDF sans page : {filename}")

    all_text = []
    with pdfplumber.open(path) as pdf:
        for index, page in enumerate(pdf.pages, start=1):
            text = (page.extract_text() or "").strip()
            # Une facture longue peut légitimement terminer par un bloc de validation,
            # de signature ou de conditions plus court. On rejette les pages réellement
            # vides, mais on autorise une dernière page courte lorsqu'elle contient un
            # marqueur métier explicite.
            if len(text) < 120:
                raise AssertionError(
                    f"Page quasi vide : {filename}, page {index}, {len(text)} caractères"
                )
            if len(text) < 250:
                is_last_page = index == len(pdf.pages)
                meaningful_last_page = is_last_page and any(
                    marker.casefold() in text.casefold()
                    for marker in (
                        "Vérification QR",
                        "Commercial Terms",
                        "TOTAL HT",
                        "SOLDE À PAYER",
                        "Informations bancaires",
                        "Signature",
                        "Préparé par",
                    )
                )
                if not meaningful_last_page:
                    raise AssertionError(
                        f"Page insuffisamment remplie : {filename}, page {index}, {len(text)} caractères"
                    )
            if any(token in text for token in ("�", "undefined", "NaN")):
                raise AssertionError(f"Valeur invalide dans {filename}, page {index}")
            all_text.append(text)

            width, height = float(page.width), float(page.height)
            portrait = close(width, PORTRAIT_A4[0]) and close(height, PORTRAIT_A4[1])
            landscape = close(width, LANDSCAPE_A4[0]) and close(height, LANDSCAPE_A4[1])
            if not (portrait or landscape):
                raise AssertionError(
                    f"Format non A4 : {filename}, page {index}, {width}x{height}"
                )

    joined = "\n".join(all_text)
    for expected in required_text:
        if expected.casefold() not in joined.casefold():
            raise AssertionError(f"Texte attendu absent de {filename} : {expected}")


xlsx = OUTPUT / "reports-sample.xlsx"
if not xlsx.exists() or xlsx.stat().st_size < 5_000:
    raise AssertionError("Export Excel absent ou anormalement petit")

workbook = load_workbook(xlsx, data_only=False)
expected_sheets = {
    "Synthese",
    "Evolution",
    "Creances",
    "Performance",
    "Bilan financier",
    "Tresorerie",
    "Dettes fournisseurs",
}
if set(workbook.sheetnames) != expected_sheets:
    raise AssertionError(f"Onglets Excel incorrects : {workbook.sheetnames}")
for sheet in workbook.worksheets:
    if sheet.max_row < 7 or sheet.max_column < 2:
        raise AssertionError(f"Onglet Excel vide ou incomplet : {sheet.title}")

print(f"PDF_AND_EXPORT_VALIDATION_OK {len(EXPECTED)} PDF + 1 XLSX")
